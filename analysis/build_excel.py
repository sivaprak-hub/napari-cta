"""
Build napari-cta Unit Testing Workbook as Excel.
Run:  python analysis/build_excel.py
Output: analysis/napari_cta_unit_testing.xlsx
"""

import os
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "napari_cta_unit_testing.xlsx")

# ── colours ───────────────────────────────────────────────────────────────────
DARK_BLUE  = "1A345C"
MID_BLUE   = "2E6DB8"
LIGHT_BLUE = "D6E8F8"
GREY_ROW   = "F4F6FA"
WHITE      = "FFFFFF"
GREEN_BG   = "E6F4EA"
GREEN_TXT  = "1A5C1A"
RED_BG     = "FDECEA"
RED_TXT    = "8B0000"
YELLOW_BG  = "FFF9E6"

def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(size=10, white=True):
    return Font(name="Calibri", bold=True, size=size,
                color=WHITE if white else DARK_BLUE)

def body_font(size=9, bold=False, color="222222"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="top", wrap_text=True)

def style_header_row(ws, row, col_count, bg=DARK_BLUE):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill    = fill(bg)
        cell.font    = hdr_font()
        cell.border  = border()
        cell.alignment = center()

def style_data_row(ws, row, col_count, alt=False):
    bg = GREY_ROW if alt else WHITE
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = fill(bg)
        cell.font      = body_font()
        cell.border    = border()
        cell.alignment = left()

def write_test_sheet(wb, sheet_name, section_intro, tests):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False

    # intro row
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = section_intro
    c.font      = Font(name="Calibri", size=9, italic=True, color="555555")
    c.fill      = fill(LIGHT_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border    = border()
    ws.row_dimensions[1].height = 28

    # column headers
    HEADERS = ["Test ID", "Test Name", "What Are We Testing",
               "How To Do It", "What Should Happen", "Pass / Fail", "Notes"]
    WIDTHS  = [9, 22, 30, 42, 35, 12, 20]

    for i, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        ws.cell(row=2, column=i).value = h

    style_header_row(ws, 2, len(HEADERS))
    ws.row_dimensions[2].height = 22

    for r_idx, t in enumerate(tests):
        row = r_idx + 3
        vals = [
            t["id"],
            t["name"],
            t["what"],
            t["steps"],
            t["expect"],
            "",
            "",
        ]
        alt = r_idx % 2 == 0
        for c_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c_idx, value=v)
        style_data_row(ws, row, len(HEADERS), alt=alt)

        # ID cell — blue bold
        id_cell = ws.cell(row=row, column=1)
        id_cell.font      = body_font(bold=True, color=MID_BLUE)
        id_cell.alignment = center()

        # Pass/Fail cell — subtle colour hint
        pf_cell = ws.cell(row=row, column=6)
        pf_cell.fill      = fill(YELLOW_BG)
        pf_cell.alignment = center()

        ws.row_dimensions[row].height = 48

    ws.freeze_panes = "A3"
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# TEST DATA  (steps = one line, plain English)
# ─────────────────────────────────────────────────────────────────────────────

FL = dict(
    name  = "File Loading",
    intro = "Can the app open the right file types and handle mistakes?",
    tests = [
        dict(id="FL-01", name="Load a TIFF file",
             what="App opens a standard .tif or .tiff image file.",
             steps="Click Add Files, pick a .tif file, click Open.",
             expect="File shows up in the queue and the image loads in the viewer."),
        dict(id="FL-02", name="Load a VSI file",
             what="App reads an Olympus VSI microscopy file.",
             steps="Click Add Files, pick a .vsi file, click Open.",
             expect="File appears in the queue and the image stack loads correctly."),
        dict(id="FL-03", name="Try a wrong file type",
             what="App shows an error — not a crash — when a non-image file is chosen.",
             steps="Click Add Files and try to open a .csv or .txt file.",
             expect="An error message appears. The app keeps running normally."),
        dict(id="FL-04", name="Add several files at once",
             what="Multiple files can be added to the queue in one go.",
             steps="Click Add Files, select 3 TIFF files using Ctrl+Click, then Open.",
             expect="All 3 files appear in the queue and can be selected one by one."),
        dict(id="FL-05", name="Remove one file",
             what="The Remove button only removes the file you have selected.",
             steps="Load 2 files, click the first one to select it, then click Remove.",
             expect="Only the first file disappears. The second stays in the queue."),
        dict(id="FL-06", name="Clear all files",
             what="Clear All removes every file and resets everything.",
             steps="Load 3 files, run analysis on one, then click Clear All and confirm.",
             expect="Queue is empty, viewer clears, graph and table are blank, verified count shows 0."),
    ]
)

FPS = dict(
    name  = "FPS & Timing",
    intro = "Is the frame rate (speed of recording) picked up correctly so that all time values make sense?",
    tests = [
        dict(id="FPS-01", name="Read FPS from file",
             what="App picks up the frame rate automatically from the file's built-in metadata.",
             steps="Load a TIFF that was saved with timing metadata and check the FPS field.",
             expect="The FPS field shows the correct value and the duration looks right (e.g. 20–120 s)."),
        dict(id="FPS-02", name="Warning when FPS is missing",
             what="If no frame rate is in the file, the app warns the user before running.",
             steps="Load a TIFF with no metadata and click Run Analysis.",
             expect="A warning dialog appears. The user can dismiss it and still run the analysis."),
        dict(id="FPS-03", name="FPS is remembered next time",
             what="If you set FPS manually, it is saved so you don't have to enter it again.",
             steps="Set FPS to 10 manually, run analysis, remove the file, then re-add the same file.",
             expect="FPS shows 10 automatically when the file is re-added. No warning pops up."),
        dict(id="FPS-04", name="Switch between FPS and Duration modes",
             what="The panel works correctly whether you enter a frame rate or a total duration.",
             steps="Switch the mode dropdown from FPS to Duration and type 60 seconds.",
             expect="The frame info label updates and the timing values are consistent."),
    ]
)

PB = dict(
    name  = "Photobleaching Correction",
    intro = "Fluorescence fades over time (photobleaching). Does the app correct for this so the signal is clean?",
    tests = [
        dict(id="PB-01", name="Single Exp correction works",
             what="The Single Exponential model removes baseline drift.",
             steps="Set the Correction Model to Single Exp and run analysis.",
             expect="Analysis finishes. Traces in the graph have flat baselines between beats, no downward drift."),
        dict(id="PB-02", name="Boundary correction works",
             what="The Boundary (valley interpolation) model also removes drift.",
             steps="Set the Correction Model to Boundary and run analysis.",
             expect="Analysis finishes without error. Baseline drift is gone from the traces."),
        dict(id="PB-03", name="Baseline sits near zero after correction",
             what="Between beats, the signal should be close to zero — not drifting.",
             steps="Run analysis, click a cell, and look at the flat parts of the trace between peaks.",
             expect="The signal stays near the zero line between beats. No obvious slope up or down."),
    ]
)

BD = dict(
    name  = "Beat Detection",
    intro = "Finding where each calcium beat starts and peaks is the foundation of all time measurements. Are the markers in the right place?",
    tests = [
        dict(id="BD-01", name="Peak markers sit on top of beats",
             what="The ▼ symbols on the graph should appear at the highest point of each beat.",
             steps="Run analysis, click on a cell with multiple clear beats, look at the ▼ markers.",
             expect="Every ▼ sits at or very close to the top of a beat. None on the flat baseline."),
        dict(id="BD-02", name="No false peaks on flat baseline",
             what="The app should not mistake noise or a flat line for a real beat.",
             steps="Click on a dim cell or an area with very little signal.",
             expect="Either no markers appear, or markers only appear where there is a real signal rise."),
        dict(id="BD-03", name="Single beat recording still works",
             what="Even if there is only one beat in the whole recording, it should be found.",
             steps="Load a short recording with only one beat, run analysis, click on an active cell.",
             expect="One ▼ marker at the peak. BPM shows a low value. T_ON and T_OFF are calculated."),
        dict(id="BD-04", name="BPM is in the right range",
             what="The reported BPM should match the known pacing frequency of the experiment.",
             steps="Run analysis on a recording paced at a known rate (e.g. 1 Hz = 60 BPM) and read the BPM column.",
             expect="BPM is within ±10% of the known rate (e.g. 54–66 for a 1 Hz recording)."),
    ]
)

KM = dict(
    name  = "Kinetic Metrics",
    intro = "These are the core scientific numbers. Every value must match the Ca²⁺ transient diagram: rise from onset, decay from peak, CD = T_ON + T_OFF.",
    tests = [
        dict(id="KM-01", name="T_ON is the rise time",
             what="T_ON_ms is the time from when the signal starts rising to when it hits the peak.",
             steps="Run analysis, click a cell with a clear beat, read T_ON_ms from the table.",
             expect="A positive number, typically 50–300 ms. It should be smaller than T_OFF_ms."),
        dict(id="KM-02", name="T_OFF is the decay time",
             what="T_OFF_ms is the time from the peak back down to the resting level.",
             steps="Run analysis, click on a cell, read T_OFF_ms from the table.",
             expect="A positive number, typically 200–800 ms. Must NOT be 3000 ms or higher."),
        dict(id="KM-03", name="Rise percentiles are in order",
             what="T10_ON < T50_ON < T90_ON — they must get progressively longer.",
             steps="Click any active cell and read T10_ON, T50_ON, T90_ON from the table.",
             expect="T10_ON is the smallest, T90_ON is the largest. All three are positive."),
        dict(id="KM-04", name="Decay percentiles are in order",
             what="T10_OFF < T50_OFF < T90_OFF — same logic for the decay.",
             steps="Click any active cell and read T10_OFF, T50_OFF, T90_OFF.",
             expect="T10_OFF is the smallest, T90_OFF is the largest. All three are positive."),
        dict(id="KM-05", name="CD = T_ON + T_OFF",
             what="Calcium Duration must equal T_ON_ms plus T_OFF_ms — nothing more, nothing less.",
             steps="Note T_ON_ms and T_OFF_ms for a cell, add them, then compare to the CD column.",
             expect="CD equals T_ON + T_OFF (e.g. 150 + 450 = 600 ms). Match within rounding."),
        dict(id="KM-06", name="Amplitude is positive",
             what="The Amp column shows how high the signal rises above the resting level.",
             steps="Click on a clearly bright, pulsing cell and read the Amp column.",
             expect="Amp is a positive number. Bright cells have higher Amp than dim cells."),
        dict(id="KM-07", name="F0 reflects resting brightness",
             what="F0 is the diastolic (resting) fluorescence — from the original uncorrected signal.",
             steps="Check the F0 column for several different cells.",
             expect="F0 values are positive and reflect how bright each cell is. Not zero or near-zero for bright cells."),
        dict(id="KM-08", name="No fake numbers for inactive cells",
             what="Cells with no real signal should show dashes, not made-up kinetic values.",
             steps="Click on a dark or inactive region of the image.",
             expect="The table row shows — for all kinetic columns (T_ON, T_OFF, CD etc)."),
    ]
)

GD = dict(
    name  = "Graph Display",
    intro = "The trace graph is how the team inspects individual cell signals. It must be readable and correctly labelled.",
    tests = [
        dict(id="GD-01", name="Graph has a white background",
             what="The plot area should always be white — never dark.",
             steps="Open the app, run analysis, click a cell, look at the graph.",
             expect="Graph background is white. Traces are colourful. Axis labels are dark and readable."),
        dict(id="GD-02", name="Each cell gets its own colour",
             what="When multiple cells are selected, each gets a unique trace colour.",
             steps="Run analysis and click on 3 different cells in the image.",
             expect="Three traces in 3 different colours. Legend shows P1, P2, P3."),
        dict(id="GD-03", name="Graph title shows the filename",
             what="You should be able to see which file you are looking at from the graph title.",
             steps="Load two files, analyse both, switch between them.",
             expect="Graph title changes to the filename of whichever file is currently active."),
        dict(id="GD-04", name="Placeholder when no cell is selected",
             what="Before clicking any cell, the graph shows a helpful message.",
             steps="Run analysis but do not click on any cell.",
             expect='Graph shows "Click on the image to select cells" in the middle.'),
        dict(id="GD-05", name="Graph stays white after theme change",
             what="Switching the napari panel theme should not turn the graph dark.",
             steps="Run analysis, then switch napari to light theme and back to dark.",
             expect="Panel colour changes but the graph canvas stays white throughout."),
    ]
)

UI = dict(
    name  = "UI State",
    intro = "When you switch between files in the queue, the app should remember what you were seeing for each file.",
    tests = [
        dict(id="UI-01", name="Results come back when you switch files",
             what="Switching away and back to a file should restore its results exactly.",
             steps="Analyse file A (note T_ON), then analyse file B, then click back on file A.",
             expect="File A's cluster map, graph, and table come back. T_ON is the same as before."),
        dict(id="UI-02", name="Verified status is not lost",
             what="Once you verify a file, it should stay verified when you switch to another file.",
             steps="Verify file A, switch to file B, then click back on file A.",
             expect="File A still shows green (verified) in the queue. The verified cell count does not drop."),
    ]
)

EX = dict(
    name  = "Excel Export",
    intro = "The final output is an Excel file with Metrics and Traces sheets. These tests make sure the export is correct.",
    tests = [
        dict(id="EX-01", name="Export creates a valid .xlsx file",
             what="The exported file must open in Excel without errors.",
             steps="Verify at least one cell, click Export master results, save, then open in Excel.",
             expect="Excel opens the file. Two sheets visible: Metrics and Traces."),
        dict(id="EX-02", name="Metrics sheet has all the right columns",
             what="Every kinetic column should be present, one row per cell.",
             steps="Open the exported file and look at the Metrics sheet.",
             expect="Columns: Filename, Cell, BPM, Amp, F0, T_ON_ms, T10_ON, T50_ON, T90_ON, T_OFF_ms, T10_OFF, T50_OFF, T90_OFF, CD."),
        dict(id="EX-03", name="CD in the export equals T_ON + T_OFF",
             what="The CD values in the spreadsheet must match the formula.",
             steps="In Excel, add a column = T_ON_ms + T_OFF_ms for each row, then compare to CD.",
             expect="New column matches the CD column for every row (within 0.1 ms)."),
        dict(id="EX-04", name="Traces sheet is laid out correctly",
             what="Time points as rows, one column per cell — makes it easy to plot in Excel.",
             steps="Open the Traces sheet in the exported file.",
             expect="First column is Time (s). Remaining columns labelled as File_P1_YxX, File_P2_YxX etc."),
        dict(id="EX-05", name="Multi-file export combines all files",
             what="Verifying cells from 2 files should produce one workbook with both.",
             steps="Verify cells from file A and file B separately, then export.",
             expect="Metrics sheet has rows from both files. Filename column shows which is which."),
    ]
)

REG = dict(
    name  = "Regression (Bug Fixes)",
    intro = "These tests specifically check that the bugs reported by the team have been fixed and do not come back.",
    tests = [
        dict(id="REG-01", name="T_OFF is never 3000 ms",
             what="The impossible 3000 ms T_OFF value that was reported must not appear anymore.",
             steps="Run analysis on a normal multi-beat recording and check T_OFF_ms for every cell.",
             expect="All T_OFF_ms values are below 1000 ms. No 3000 ms anywhere."),
        dict(id="REG-02", name="Peak markers are on peaks, not the baseline",
             what="The ▼ markers must only appear at real calcium peaks.",
             steps="Run analysis and click through 5–10 different cells, checking each trace.",
             expect="Every ▼ is at the top of a real beat. None appear on the flat parts between beats."),
        dict(id="REG-03", name="CD = T_ON + T_OFF (confirmed)",
             what="This is a direct check of the fixed formula.",
             steps="Pick any cell and manually add T_ON_ms + T_OFF_ms, then compare to CD.",
             expect="They match. For example: 150 + 450 = 600 ms CD."),
        dict(id="REG-04", name="Graph background is white",
             what="The graph must have a white background — dark background was a reported bug.",
             steps="Open the app in default dark mode, run analysis, click a cell.",
             expect="Graph area is white. No dark background inside the plot."),
        dict(id="REG-05", name="Clear All button works",
             what="The Clear All button was missing and has been added — confirm it is there and works.",
             steps="Load 2 files, find the Clear All button next to Remove, click it and confirm.",
             expect="All files removed, viewer cleared, verified count resets to 0."),
    ]
)

ALL_SECTIONS = [FL, FPS, PB, BD, KM, GD, UI, EX, REG]

# ─────────────────────────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default Sheet

# ── OVERVIEW sheet ─────────────────────────────────────────────────────────────
ws_ov = wb.create_sheet("Overview")
ws_ov.sheet_view.showGridLines = False

TIMELINE = [
    ("v0.1", "16 Jun 2026", "First working version: image loading, photobleaching correction, clustering, kinetics, Excel export."),
    ("v0.2", "09 Jul 2026", "Science overhaul: FPS sidecar, exponential decay fit, per-beat averaging, CD = T_ON + T_OFF."),
    ("v0.3", "19 Jul 2026", "Per-file UI state: switching files now restores the graph, table, and cluster map."),
    ("v0.4", "24 Jul 2026", "Raw signal F0 fix, full dark theme, detachable graph window, valley-clipped decay window."),
    ("v0.5", "29 Jul 2026", "Bug fix sprint: T_OFF 3000 ms fixed, peak markers fixed, white graph background, Clear All button added."),
    ("v0.6", "31 Jul 2026", "Light/dark theme toggle: panel responds to napari theme switch. Graph always stays white."),
]

WHAT_TESTED = [
    ("File Loading",         "Can the app open TIFF and VSI files? Does it handle bad inputs?",                "FL"),
    ("FPS & Timing",         "Is the frame rate detected correctly? Are time values in sensible units?",       "FPS"),
    ("Photobleaching",       "Is the baseline drift (bleaching) removed from the fluorescence traces?",       "PB"),
    ("Beat Detection",       "Are calcium peaks found in the right places? No false positives?",              "BD"),
    ("Kinetic Metrics",      "Are T_ON, T_OFF, CD and percentile times biologically correct?",                "KM"),
    ("Graph Display",        "Is the trace graph readable — white, correct colours, correct labels?",         "GD"),
    ("UI State",             "Does switching between files restore results without losing anything?",          "UI"),
    ("Excel Export",         "Is the output file correct, complete, and can Excel open it?",                  "EX"),
    ("Regression (Bugs)",    "Do the specific bugs reported by the team stay fixed?",                         "REG"),
]

# Title
ws_ov.merge_cells("A1:D1")
t = ws_ov["A1"]
t.value     = "napari-CTA  —  Calcium Transient Analyser  |  Unit Testing"
t.font      = Font(name="Calibri", bold=True, size=16, color=WHITE)
t.fill      = fill(DARK_BLUE)
t.alignment = center()
ws_ov.row_dimensions[1].height = 36

ws_ov.merge_cells("A2:D2")
sub = ws_ov["A2"]
sub.value     = "Repository: sivaprak-hub/napari-cta     |     Use the sheet tabs below to run each test section"
sub.font      = Font(name="Calibri", size=9, italic=True, color="555555")
sub.fill      = fill(LIGHT_BLUE)
sub.alignment = center()
ws_ov.row_dimensions[2].height = 18

ws_ov.row_dimensions[3].height = 10   # spacer

# Version timeline
ws_ov["A4"].value = "Version History"
ws_ov["A4"].font  = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
ws_ov.row_dimensions[4].height = 20

tl_headers = ["Version", "Date", "What Changed"]
tl_widths   = [10, 15, 80]
for i, (h, w) in enumerate(zip(tl_headers, tl_widths), 1):
    ws_ov.column_dimensions[get_column_letter(i)].width = w
    c = ws_ov.cell(row=5, column=i, value=h)
    c.fill = fill(MID_BLUE); c.font = hdr_font(9); c.alignment = center(); c.border = border()
ws_ov.row_dimensions[5].height = 18

for r, (ver, date, desc) in enumerate(TIMELINE, 6):
    alt = (r % 2 == 0)
    bg  = GREY_ROW if alt else WHITE
    for col, v in enumerate([ver, date, desc], 1):
        c = ws_ov.cell(row=r, column=col, value=v)
        c.fill = fill(bg); c.border = border()
        c.alignment = Alignment(horizontal="left" if col == 3 else "center",
                                 vertical="center", wrap_text=True)
        c.font = body_font(9, bold=(col == 1), color=MID_BLUE if col == 1 else "222222")
    ws_ov.row_dimensions[r].height = 22

ws_ov.row_dimensions[r + 1].height = 12   # spacer

# What is being tested
ws_ov.cell(row=r+2, column=1).value = "What We Are Testing"
ws_ov.cell(row=r+2, column=1).font  = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
ws_ov.row_dimensions[r+2].height = 20

wt_hdr_row = r + 3
wt_headers = ["Area", "What We Check", "Sheet"]
wt_widths  = [10, 80, 10]  # A B and a 4th col for sheet name
for i, (h, w) in enumerate(zip(wt_headers, wt_widths), 1):
    c = ws_ov.cell(row=wt_hdr_row, column=i, value=h)
    c.fill = fill(MID_BLUE); c.font = hdr_font(9); c.alignment = center(); c.border = border()
ws_ov.row_dimensions[wt_hdr_row].height = 18

for rr, (area, desc, sheet) in enumerate(WHAT_TESTED, wt_hdr_row + 1):
    alt = (rr % 2 == 0)
    bg  = GREY_ROW if alt else WHITE
    for col, v in enumerate([area, desc, sheet], 1):
        c = ws_ov.cell(row=rr, column=col, value=v)
        c.fill = fill(bg); c.border = border()
        c.font = body_font(9, bold=(col == 1), color=MID_BLUE if col == 1 else "222222")
        c.alignment = Alignment(horizontal="left" if col == 2 else "center",
                                 vertical="center", wrap_text=True)
    ws_ov.row_dimensions[rr].height = 20

# fix column D width
ws_ov.column_dimensions["D"].width = 10

# ── TEST SECTIONS ──────────────────────────────────────────────────────────────
for sec in ALL_SECTIONS:
    write_test_sheet(wb, sec["name"], sec["intro"], sec["tests"])

# ── SIGN-OFF sheet ─────────────────────────────────────────────────────────────
ws_so = wb.create_sheet("Sign-Off")
ws_so.sheet_view.showGridLines = False
ws_so.column_dimensions["A"].width = 22
ws_so.column_dimensions["B"].width = 50

ws_so.merge_cells("A1:B1")
c = ws_so["A1"]
c.value = "Testing Sign-Off"
c.font  = Font(name="Calibri", bold=True, size=14, color=WHITE)
c.fill  = fill(DARK_BLUE); c.alignment = center()
ws_so.row_dimensions[1].height = 30

FIELDS = [
    ("Tester Name",          ""),
    ("Date of Testing",      ""),
    ("Version Tested",       ""),
    ("Overall Result",       "Pass  /  Fail  (delete one)"),
    ("Tests That Failed",    "e.g. KM-05, REG-01"),
    ("Notes / Observations", ""),
    ("Signature",            ""),
]

for r, (label, hint) in enumerate(FIELDS, 2):
    alt = r % 2 == 0
    bg  = GREY_ROW if alt else WHITE
    la = ws_so.cell(row=r, column=1, value=label)
    la.font = body_font(10, bold=True, color=DARK_BLUE)
    la.fill = fill(bg); la.border = border()
    la.alignment = Alignment(horizontal="left", vertical="center")

    va = ws_so.cell(row=r, column=2, value=hint)
    va.font = body_font(10, color="888888")
    va.fill = fill(bg); va.border = border()
    va.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws_so.row_dimensions[r].height = 28 if label != "Signature" else 50

# ── SAVE ───────────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"Saved: {OUT}")
